import openpyxl as xl
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Font, Alignment
import pandas as pd

def createExcel(wbName, headers):
    wb = xl.Workbook()
    page = wb.active
    page.title = 'Simulation results'
    page.append(headers)
    page.freeze_panes = page['a2']
    page.column_dimensions['j'].width = 70
    page.column_dimensions['k'].width = 50
    page.column_dimensions['l'].width = 80
    page.column_dimensions['C'].width = 16
    page.column_dimensions['D'].width = 13
    page.column_dimensions['E'].width = 13
    page['A1'].font = Font(bold=True)
    page['B1'].font = Font(bold=True)
    page['C1'].font = Font(bold=True)
    page['D1'].font = Font(bold=True)
    page['E1'].font = Font(bold=True)
    page['F1'].font = Font(bold=True)
    page['G1'].font = Font(bold=True)
    page['H1'].font = Font(bold=True)
    page['I1'].font = Font(bold=True)
    page['J1'].font = Font(bold=True)
    page['K1'].font = Font(bold=True)
    page['L1'].font = Font(bold=True)
    wb.save(wbName)

def writeExcel(wbName, parameters, barPlt, energyPlt, timePlt):
    wb = xl.load_workbook(wbName)
    page = wb.active
    page.append(parameters)
    numrows = page.max_row
    page.row_dimensions[numrows].height= 240
    page['A'+str(numrows)].alignment = Alignment(horizontal='center', vertical='center')
    page['B'+str(numrows)].alignment = Alignment(horizontal='center', vertical='center')
    page['C'+str(numrows)].alignment = Alignment(horizontal='center', vertical='center')
    page['D'+str(numrows)].alignment = Alignment(horizontal='center', vertical='center')
    page['E'+str(numrows)].alignment = Alignment(horizontal='center', vertical='center')
    page['F'+str(numrows)].alignment = Alignment(horizontal='center', vertical='center')
    page['G'+str(numrows)].alignment = Alignment(horizontal='center', vertical='center')
    page['H'+str(numrows)].alignment = Alignment(horizontal='center', vertical='center')
    imgBar = xl.drawing.image.Image(barPlt)
    imgBar.anchor = TwoCellAnchor(editAs='twoCell', _from= AnchorMarker(col=9, row=str(numrows-1)), to=AnchorMarker(col=10, row=str(numrows)))
    page.add_image(imgBar)
    imgEnergy = xl.drawing.image.Image(energyPlt)
    imgEnergy.anchor = TwoCellAnchor(editAs='twoCell', _from= AnchorMarker(col=10, row=str(numrows-1)), to=AnchorMarker(col=11, row=str(numrows)))
    page.add_image(imgEnergy)
    imgTime = xl.drawing.image.Image(timePlt)
    imgTime.anchor = TwoCellAnchor(editAs='twoCell', _from= AnchorMarker(col=11, row=str(numrows-1)), to=AnchorMarker(col=12, row=str(numrows)))
    page.add_image(imgTime)
    wb.save(wbName)